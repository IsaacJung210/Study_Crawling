import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임')

# 3) 데이터 추가하기
ws['A1'] = '종목'
ws['B1'] = '현재가'
ws['C1'] = '평균매입가'
ws['D1'] = '잔고수량'
ws['E1'] = '평가금액'
ws['F1'] = '평가손익'
ws['G1'] = '수익률'


# 4) 엑셀 저장하기 (r은 이스케이프 문자를 그냥 문자로 인식시킴 or \\ or / 사용)
wb.save(r'C:\Users\user\Desktop\인프런\웹크롤링\STARTCODING_CRAWLING\01_네이버_주식현재가_크롤링\data.xlsx')