import openpyxl

# 경로 정하기 (경로복사하면 쉬움)
fpath = r'C:\Users\user\Desktop\인프런\웹크롤링\STARTCODING_CRAWLING\02_파이썬엑셀다루기\참가자_data.xlsx'

# 1) 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

# 2) 엑셀 시트선택
ws = wb['오징어게임']

# 3) 데이터 수정하기
ws['A3'] = 456
ws['b3'] = '성기훈'

# 4) 엑셀 저장하기
wb.save(fpath)


#############################공식문서####################################
## https://openpyxl.readthedocs.io/en/stable/tutorial.html
########################################################################
