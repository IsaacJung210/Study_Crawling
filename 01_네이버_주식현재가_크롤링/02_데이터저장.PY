import requests
from bs4 import BeautifulSoup
import openpyxl

fpath = r'C:\Users\user\Desktop\인프런\웹크롤링\STARTCODING_CRAWLING\01_네이버_주식현재가_크롤링\data.xlsx'
wb = openpyxl.load_workbook(fpath)
ws = wb.active # 현재 활성화된 시트 선택

# 종목 코드 리스트
codes = [
    '005930',
    '000660',
    '035720'
]

row = 2
for code in codes:
    url = f"https://finance.naver.com/item/sise.naver?code={code}"
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    title = soup.select_one('.wrap_company > h2 > a').text
    price = soup.select_one('#_nowVal').text
    price = price.replace(',','')
    print(title, price)
    ws[f'A{row}'] = str(title)
    ws[f'B{row}'] = int(price)
    row = row + 1

wb.save(fpath)
